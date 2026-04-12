# Copyright 2025 Softwell S.r.l. - Licensed under Apache License 2.0
# See LICENSE file for details

"""ExcelCompiler - Compiler for Excel spreadsheets (.xlsx).

Transforms a built Bag (with ExcelBuilder) into an Excel workbook using
the genro-builders @compiler dispatch and auto-walk. Handlers receive
the node and parent live object. Resolved attributes are accessed via
node.runtime_attrs, resolved value via node.runtime_value.
"""

from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING, Any

from genro_builders import BagCompilerBase
from genro_builders.compiler import compiler
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

if TYPE_CHECKING:
    from genro_bag import Bag, BagNode


class ExcelCompiler(BagCompilerBase):
    """Compiler Excel: Bag -> bytes (.xlsx).

    Uses @compiler handlers with parent propagation.
    Maintains a live Workbook and _live_map for incremental updates.
    """

    def __init__(self, builder: Any) -> None:
        super().__init__(builder)
        self._wb: Any | None = None
        self._live_map: dict[int, Any] = {}
        self._custom_handlers: dict[str, Any] = {}
        self._current_row_idx: int = 0
        self._current_col_idx: int = 0

    def register_handler(self, tag: str, handler: Any) -> None:
        """Register a custom compile handler for a tag.

        The handler receives (node, parent) and is called during workbook
        building. Custom handlers take precedence over @compiler ones.

        Args:
            tag: The element tag name (e.g., "sparkline").
            handler: Callable(node, parent).
        """
        self._custom_handlers[tag] = handler

    def _dispatch_compile(self, node: BagNode, parent: Any = None) -> Any | None:
        """Override dispatch to check custom handlers first."""
        tag = node.node_tag or node.label

        custom = self._custom_handlers.get(tag)
        if custom is not None:
            return custom(node, parent)

        return super()._dispatch_compile(node, parent=parent)

    # -------------------------------------------------------------------------
    # Compile / Render / Serialize
    # -------------------------------------------------------------------------

    def compile(self, built_bag: Bag, target: Any = None) -> Any:  # noqa: ARG002
        """Compile built Bag into a live Workbook."""
        self._wb = Workbook()
        self._live_map.clear()

        if self._wb.active:
            self._wb.remove(self._wb.active)

        list(self._walk_compile(built_bag, parent=self._wb))

        if not self._wb.sheetnames:
            self._wb.create_sheet("Sheet1")

        return self._wb

    def render(self, compiled_bag: Bag) -> bytes:
        """Compile and serialize to bytes (.xlsx)."""
        self.compile(compiled_bag)
        return self.serialize()

    def serialize(self) -> bytes:
        """Serialize the live Workbook to bytes without rebuilding."""
        if self._wb is None:
            return b""
        buffer = BytesIO()
        self._wb.save(buffer)
        return buffer.getvalue()

    # -------------------------------------------------------------------------
    # Live update
    # -------------------------------------------------------------------------

    def update_node(self, node: BagNode) -> bool:
        """Try to update the live object for a node."""
        live_obj = self._live_map.get(id(node))
        if live_obj is None:
            return False
        attrs = node.runtime_attrs
        return self._apply_live_update(attrs, node.node_tag or "", live_obj)

    def _apply_live_update(
        self, attrs: dict[str, Any], tag: str, live_obj: Any
    ) -> bool:
        """Apply resolved attribute changes to a live openpyxl object."""
        if tag == "cell":
            formula = attrs.get("formula")
            if formula:
                live_obj.value = formula
            else:
                live_obj.value = attrs.get("content", "")

            self._apply_font(attrs, live_obj)
            self._apply_alignment(attrs, live_obj)
            self._apply_border(attrs, live_obj)

            bg_color = attrs.get("bg_color")
            if bg_color:
                live_obj.fill = PatternFill(
                    start_color=str(bg_color), end_color=str(bg_color), fill_type="solid"
                )

            number_format = attrs.get("number_format")
            if number_format:
                live_obj.number_format = str(number_format)

            return True

        if tag == "row":
            height = attrs.get("height")
            if height is not None:
                live_obj.height = height
            hidden = attrs.get("hidden", False)
            live_obj.hidden = bool(hidden)
            return True

        return False

    # -------------------------------------------------------------------------
    # @compiler handlers — auto-walked by _dispatch_compile
    # -------------------------------------------------------------------------

    @compiler()
    def workbook(self, node: BagNode, parent: Any) -> Any:  # noqa: ARG002
        """Compile workbook node. Parent is the Workbook."""
        return parent

    @compiler()
    def sheet(self, node: BagNode, parent: Any) -> Any:
        """Compile sheet node. Parent is the Workbook."""
        attrs = node.runtime_attrs
        name = attrs.get("name", "Sheet1")
        ws = parent.create_sheet(title=str(name))

        freeze_panes = attrs.get("freeze_panes")
        if freeze_panes:
            ws.freeze_panes = str(freeze_panes)

        autofilter = attrs.get("autofilter")
        if autofilter:
            ws.auto_filter.ref = str(autofilter)

        self._current_row_idx = 0
        return ws

    @compiler()
    def row(self, node: BagNode, parent: Any) -> Any:
        """Compile row node. Parent is the Worksheet."""
        attrs = node.runtime_attrs
        ws = parent
        self._current_row_idx += 1
        row_idx = self._current_row_idx

        row_dim = ws.row_dimensions[row_idx]
        self._live_map[id(node)] = row_dim

        height = attrs.get("height")
        if height:
            row_dim.height = height

        hidden = attrs.get("hidden", False)
        if hidden:
            row_dim.hidden = True

        self._current_col_idx = 0
        return ws

    @compiler()
    def cell(self, node: BagNode, parent: Any) -> Any:
        """Compile cell node. Parent is the Worksheet (passed through from row)."""
        attrs = node.runtime_attrs
        ws = parent
        self._current_col_idx += 1
        row_idx = self._current_row_idx
        col_idx = self._current_col_idx

        content = attrs.get("content", "")
        formula = attrs.get("formula")
        width = attrs.get("width")

        xl_cell = ws.cell(row=row_idx, column=col_idx)
        self._live_map[id(node)] = xl_cell

        if formula:
            xl_cell.value = formula
        else:
            xl_cell.value = content

        if width:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        self._apply_font(attrs, xl_cell)

        bg_color = attrs.get("bg_color")
        if bg_color:
            xl_cell.fill = PatternFill(
                start_color=str(bg_color), end_color=str(bg_color), fill_type="solid"
            )

        self._apply_alignment(attrs, xl_cell)
        self._apply_border(attrs, xl_cell)

        number_format = attrs.get("number_format")
        if number_format:
            xl_cell.number_format = str(number_format)

        return xl_cell

    @compiler()
    def merge(self, node: BagNode, parent: Any) -> Any:
        """Compile merge node. Parent is the Worksheet."""
        attrs = node.runtime_attrs
        cell_range = attrs.get("range", "")
        if cell_range:
            parent.merge_cells(str(cell_range))
        return None

    @compiler()
    def chart(self, node: BagNode, parent: Any) -> Any:
        """Compile chart node. Parent is the Worksheet."""
        attrs = node.runtime_attrs
        ws = parent
        chart_type = attrs.get("type", "bar")
        title = attrs.get("title")
        data_range = attrs.get("data_range", "")
        categories_range = attrs.get("categories_range")
        position = attrs.get("position", "E1")
        width = attrs.get("width", 15)
        height = attrs.get("height", 10)

        if not data_range:
            return None

        if chart_type == "line":
            chart_obj = LineChart()
        elif chart_type == "pie":
            chart_obj = PieChart()
        else:
            chart_obj = BarChart()

        if title:
            chart_obj.title = str(title)

        chart_obj.width = width
        chart_obj.height = height

        data_ref = self._parse_range_reference(ws, str(data_range))
        if data_ref:
            chart_obj.add_data(data_ref, titles_from_data=True)

        if categories_range:
            cat_ref = self._parse_range_reference(ws, str(categories_range))
            if cat_ref:
                chart_obj.set_categories(cat_ref)

        ws.add_chart(chart_obj, str(position))
        return None

    def _parse_range_reference(self, ws: Any, range_str: str) -> Any:
        """Parse a range string into a Reference object."""
        if ":" not in range_str:
            return None

        start, end = range_str.split(":")

        start_col = ""
        start_row = ""
        for char in start:
            if char.isalpha():
                start_col += char
            else:
                start_row += char

        end_col = ""
        end_row = ""
        for char in end:
            if char.isalpha():
                end_col += char
            else:
                end_row += char

        min_col = column_index_from_string(start_col)
        max_col = column_index_from_string(end_col)
        min_row = int(start_row)
        max_row = int(end_row)

        return Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)

    # -------------------------------------------------------------------------
    # Formatting helpers
    # -------------------------------------------------------------------------

    def _apply_font(self, attrs: dict[str, Any], cell: Any) -> None:
        """Apply font styling to cell."""
        bold = attrs.get("bold", False)
        italic = attrs.get("italic", False)
        underline = attrs.get("underline", False)
        font_size = attrs.get("font_size")
        font_color = attrs.get("font_color")

        if bold or italic or underline or font_size or font_color:
            cell.font = Font(
                bold=bold,
                italic=italic,
                underline="single" if underline else None,
                size=font_size,
                color=str(font_color) if font_color else None,
            )

    def _apply_alignment(self, attrs: dict[str, Any], cell: Any) -> None:
        """Apply alignment to cell."""
        align = attrs.get("align")
        valign = attrs.get("valign")
        wrap_text = attrs.get("wrap_text", False)

        if align or valign or wrap_text:
            cell.alignment = Alignment(
                horizontal=str(align) if align else None,
                vertical=str(valign) if valign else None,
                wrap_text=wrap_text,
            )

    def _apply_border(self, attrs: dict[str, Any], cell: Any) -> None:
        """Apply border to cell."""
        border_style = attrs.get("border")
        border_color = attrs.get("border_color", "000000")

        if border_style:
            side = Side(style=str(border_style), color=str(border_color))
            cell.border = Border(left=side, right=side, top=side, bottom=side)


# Set _compiler_class after ExcelCompiler is defined
from genro_office.builders.excel_builder import ExcelBuilder  # noqa: E402

ExcelBuilder._compiler_class = ExcelCompiler
